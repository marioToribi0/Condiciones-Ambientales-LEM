from csv import excel
import pandas as pd
import openpyxl

from datetime import datetime, timedelta
import pytz

from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

import smtplib,ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
 
# Object Border
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
 
def generar_reporte(csv, number: str, name: str):
    data = pd.read_csv(csv, sep=";", header=3, encoding='unicode_escape')
    data.columns = ['SN', 'DATE', 'TIME', 'oC', '%RH', 'DP']
    def cleaned_date(x):
        values = x.split("/")
        return f"{int(values[0]):02d}/{int(values[1]):02d}/{int(values[2])}"

    def cleaned_time(x):
        values = x.split(":")
        return f"{int(values[0]):02d}:{int(values[1]):02d}:{int(values[2]):02d}"

    data.DATE = data.DATE.apply(lambda x: cleaned_date(x))
    data.TIME = data.TIME.apply(lambda x: cleaned_time(x))
    data.TIME = pd.to_datetime(data.DATE + ' ' +data.TIME, format='%d/%m/%Y %H:%M:%S')

    years = data.TIME.dt.year.value_counts()
    years = years.index.to_list()

    data = data.drop('DATE', axis=1)

    hour_9_to_1 = data[(data.TIME.dt.hour>=9) & (data.TIME.dt.hour<13)]
    hour_2_to_6 = data[(data.TIME.dt.hour>=14) & (data.TIME.dt.hour<18)]
    hour_7_to_10 = data[(data.TIME.dt.hour>=19) & (data.TIME.dt.hour<22)]

    # Monday to Sunday
    hour_2_to_6 = hour_2_to_6[hour_2_to_6.TIME.dt.dayofweek!=6]
    hour_9_to_1 = hour_9_to_1[hour_9_to_1.TIME.dt.dayofweek!=6]
    hour_7_to_10 = hour_7_to_10[hour_7_to_10.TIME.dt.dayofweek!=6]

    max_month = hour_2_to_6.groupby([hour_2_to_6.TIME.dt.month]).count()['TIME']
    max_month = max_month[max_month.values==max_month.max()].index[0]

    hour_9_to_1 = hour_9_to_1[hour_9_to_1.TIME.dt.month==max_month]
    hour_2_to_6 = hour_2_to_6[hour_2_to_6.TIME.dt.month==max_month]
    hour_7_to_10 = hour_7_to_10[hour_7_to_10.TIME.dt.month==max_month]

    hour_9_to_1 = hour_9_to_1.groupby([hour_9_to_1.TIME.dt.date]).apply(lambda x: x.sample(1))
    hour_2_to_6 = hour_2_to_6.groupby([hour_2_to_6.TIME.dt.date]).apply(lambda x: x.sample(1))
    hour_7_to_10 = hour_7_to_10.groupby([hour_7_to_10.TIME.dt.date]).apply(lambda x: x.sample(1))

    hour_7_to_10["TIME"] = list(map(lambda x: x[0], hour_7_to_10.index.to_list()))
    hour_2_to_6["TIME"] = list(map(lambda x: x[0], hour_2_to_6.index.to_list()))
    hour_9_to_1["TIME"] = list(map(lambda x: x[0], hour_9_to_1.index.to_list()))

    data = [hour_9_to_1, hour_7_to_10, hour_2_to_6]
    for el in data:
        el.drop(["DP","SN"], axis=1, inplace=True)
        
    days = []
    months_name = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    year = years[0]
    
    # Open base excel
    path = f"./static/static"
    wb_obj = openpyxl.load_workbook(f"{path}/base.xlsx") 
    #wb_obj = openpyxl.load_workbook(f"Base1.xlsx") 
    month = months_name[max_month-1]

    sheet_obj = wb_obj['Mes']

    sheet_obj.cell(row=7, column=2).value = month
    
    bad_temperature = 0
    bad_humedity = 0
    
    labor_days = len(hour_2_to_6)

    for b in range(1, 4):
        df = data[b-1]
        day = list(filter(lambda x: x.month==max_month, map(lambda x: x[0], df.index.to_list())))
        # First update days
        days = list(map(lambda x: x.day, day))
        
        temperature = df[pd.to_datetime(df["TIME"]).dt.month==max_month]["oC"].to_list()
        humedity = df[pd.to_datetime(df["TIME"]).dt.month==max_month]["%RH"].to_list()
        
        bad_temperature += sum([1 if t<10 or t>=35 else 0 for t in temperature])
        bad_humedity += sum([1 if t<10 or t>=80 else 0 for t in humedity])
        
        # Add data to excel
        for i, day in enumerate(days):
            # Temperature
            cell_obj = sheet_obj.cell(row = day+10, column = b+1)
            cell_obj.value =  temperature[i]
            
            
            # Humedity
            cell_obj = sheet_obj.cell(row = day+10, column = b+21)
            cell_obj.value =  humedity[i]

    # Agregar en reporte lista de dias
    
    mean_temperature = []
    for i in range(len(hour_9_to_1['oC'])):
        actual = 0
        n = 1
        actual += hour_9_to_1['oC'][i]
        if (len(hour_2_to_6["oC"])==len(hour_9_to_1) and i==len(hour_9_to_1)-1):
            n += 1
            actual += hour_2_to_6['oC'][i]
        if (len(hour_7_to_10["oC"])==len(hour_9_to_1) and i==len(hour_9_to_1)-1):
            n += 1
            actual += hour_7_to_10['oC'][i]
        
        mean_temperature.append(actual/n)
    mean_humedity = []
    for i in range(len(hour_9_to_1['%RH'])):
        actual = 0
        n = 1
        actual += hour_9_to_1['%RH'][i]
        if (len(hour_2_to_6["%RH"])==len(hour_9_to_1) and i==len(hour_9_to_1)-1):
            n += 1
            actual += hour_2_to_6['%RH'][i]
        if (len(hour_7_to_10["%RH"])==len(hour_9_to_1) and i==len(hour_9_to_1)-1):
            n += 1
            actual += hour_7_to_10['%RH'][i]
        
        mean_humedity.append(actual/n)

    
    resume = pd.DataFrame({"TIME": hour_9_to_1["TIME"], "TEMPERATURE":mean_temperature,
                       "HUMEDITY": mean_humedity})
    sheet_obj = wb_obj['Reporte']

    for i in range(len(resume["TIME"])):
        date = f"{resume['TIME'][i].year}-{int(resume['TIME'][i].month):02d}-{int(resume['TIME'][i].day):02d}"
        humedity = resume["HUMEDITY"][i]
        temperature = resume["TEMPERATURE"][i]
        
        sheet_obj.cell(row=9+i, column=1).value = date
        sheet_obj.cell(row=9+i, column=1).border = thin_border
        
        sheet_obj.cell(row=9+i, column=2).value = temperature
        sheet_obj.cell(row=9+i, column=2).border = thin_border
        
        sheet_obj.cell(row=9+i, column=3).value = humedity
        sheet_obj.cell(row=9+i, column=3).border = thin_border
    # Mensaje
    data[0]["TANDA"] = "mañana"
    data[1]["TANDA"] = "tarde"
    data[2]["TANDA"] = "noche"
    concat_data = pd.concat(data)
    min_temperature = concat_data[concat_data["oC"]==concat_data["oC"].min()]
    max_temperature = concat_data[concat_data["oC"]==concat_data["oC"].max()]
    min_humedity = concat_data[concat_data["%RH"]==concat_data["%RH"].min()]
    max_humedity = concat_data[concat_data["%RH"]==concat_data["%RH"].max()]
    # Add data
    sheet_obj.cell(row=39, column=11).value = max_temperature["oC"][0]
    sheet_obj.cell(row=39, column=12).value = max_temperature["TANDA"][0]
    sheet_obj.cell(row=39, column=13).value = f"{int(max_temperature['TIME'][0].day):02d}"
    #
    sheet_obj.cell(row=40, column=11).value = min_temperature["oC"][0]
    sheet_obj.cell(row=40, column=12).value = min_temperature["TANDA"][0]
    sheet_obj.cell(row=40, column=13).value = f"{int(min_temperature['TIME'][0].day):02d}"
    #
    sheet_obj.cell(row=41, column=11).value = max_humedity["%RH"][0]
    sheet_obj.cell(row=41, column=12).value = max_humedity["TANDA"][0]
    sheet_obj.cell(row=41, column=13).value = f"{int(max_humedity['TIME'][0].day):02d}"
    #
    sheet_obj.cell(row=42, column=11).value = min_humedity["%RH"][0]
    sheet_obj.cell(row=42, column=12).value = min_humedity["TANDA"][0]
    sheet_obj.cell(row=42, column=13).value = f"{int(min_humedity['TIME'][0].day):02d}"
    #
    sheet_obj.cell(row=43, column=11).value = bad_humedity
    sheet_obj.cell(row=43, column=11).value = bad_temperature
    
    # Correlativo
    sheet_obj.cell(row=4, column=10).value = number
    sheet_obj.cell(row=5, column=2).value = year
    
    # Nombre
    sheet_obj["C40"] = name
    
    # Fecha
    UTC = pytz.utc
    now = datetime.now(UTC) - timedelta(hours=4)

    sheet_obj["C41"] = f"{now.year}-{now.month}-{now.day:02d}"
    
    bad_temperature = concat_data.query("oC>=35 or oC<=10")
    concat_data["RH"] = concat_data["%RH"]
    bad_humedity = concat_data.query("RH>=80 or RH<=10")


    message = f'''Temperatura máxima: {max_temperature["oC"][0]}ºC {max_temperature["TANDA"][0]} del {int(max_temperature['TIME'][0].day):02d}\nTemperatura mínima: {min_temperature["oC"][0]}ºC {min_temperature["TANDA"][0]} del {int(min_temperature['TIME'][0].day):02d}\nHumedad máxima: {max_humedity["%RH"][0]}ºC {max_humedity["TANDA"][0]} del {int(max_humedity['TIME'][0].day):02d}\nHumedad mínima: {min_humedity["%RH"][0]}ºC {min_humedity["TANDA"][0]} del {int(min_humedity['TIME'][0].day):02d}\n\n'''
    message_2 = ""
    
    if (len(bad_temperature)==0):
        message += f"Temperatura fuera de los límites: {len(bad_temperature)}\n"
        message_2 += f"Temperatura fuera de los límites: {len(bad_temperature)}\n"
    else:
        message += f"Temperatura fuera de los límites: {len(bad_temperature)}\n"
        message_2 += f"Temperatura fuera de los límites: {len(bad_temperature)} ("
        
        for i in range(len(bad_temperature)):
            message += f"\t->{bad_temperature['oC'][i]}ºC {bad_temperature['TANDA'][i]} del {int(bad_temperature['TIME'][i].day):02d}\n"
            message_2 += f"{bad_temperature['oC'][i]}ºC {int(bad_temperature['TIME'][i].day):02d} {bad_temperature['TANDA'][i]}"
            
            if (i!=len(bad_temperature)-1):
                message_2 += f", "
            else:
                message_2 += ")\n"
                
            
    if (len(bad_humedity)==0):
        message += f"Humedad fuera de los límites: {len(bad_humedity)}\n"
        message_2 += f"Humedad fuera de los límites: {len(bad_humedity)}\n"
    else:
        message += f"Humedad fuera de los límites: {len(bad_humedity)}\n"
        message_2 += f"Humedad fuera de los límites: {len(bad_humedity)} ("
        for i in range(len(bad_humedity)):
            message += f"\t->{bad_humedity['%RH'][i]}% {bad_humedity['TANDA'][i]} del {int(bad_humedity['TIME'][i].day):02d}\n"
            
            message_2 += f"{bad_humedity['%RH'][i]}% {int(bad_humedity['TIME'][i].day):02d} {bad_humedity['TANDA'][i]}"
            
            if (i!=len(bad_humedity)-1):
                message_2 += f", "
            else:
                message_2 += ")\n"
    
    sheet_obj["A45"] = message_2
    excel = f"./static/outputs/LEM-F-6.3-01-08 Registro de Condiciones Ambientales v.1 {number}.xlsx"
    wb_obj.save(excel)
    
    return excel, labor_days, year, month, message

def send_mail(send_from,send_to,subject,text,excel,server,port,username='',password='',isTls=True):
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = send_to
    msg['Date'] = formatdate(localtime = True)
    msg['Subject'] = subject
    msg.attach(MIMEText(text))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(excel, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{excel.split("/")[-1]}"')
    msg.attach(part)

    #SSL connection only working on Python 3+
    smtp = smtplib.SMTP(server, port)
    if isTls:
        smtp.starttls()

    smtp.login(username,password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()
